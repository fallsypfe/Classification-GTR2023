import streamlit as st
from datetime import datetime
from io import BytesIO
from pathlib import Path
import base64, json

st.set_page_config(page_title="Classification GTR2023", page_icon="⏳", layout="centered")

LOGO_ECOLE = Path(__file__).parent / "logo_ecole.png"
def img_b64(p):
    return base64.b64encode(p.read_bytes()).decode() if p.exists() else ""

LOGO_SVG = """<svg viewBox="0 0 200 150" xmlns="http://www.w3.org/2000/svg" width="80" style="display:block;margin:0 auto;">
<rect x="30" y="45" width="140" height="16" rx="4" fill="#8a4a14"/><rect x="30" y="57" width="140" height="4" rx="1" fill="#b5651d"/>
<line x1="55" y1="61" x2="55" y2="57" stroke="#6e5847" stroke-width="1.5"/><line x1="75" y1="61" x2="75" y2="57" stroke="#6e5847" stroke-width="1.5"/>
<line x1="95" y1="61" x2="95" y2="57" stroke="#6e5847" stroke-width="1.5"/><line x1="115" y1="61" x2="115" y2="57" stroke="#6e5847" stroke-width="1.5"/>
<line x1="135" y1="61" x2="135" y2="57" stroke="#6e5847" stroke-width="1.5"/><line x1="65" y1="61" x2="65" y2="57" stroke="#6e5847" stroke-width="1.5"/>
<line x1="85" y1="61" x2="85" y2="57" stroke="#6e5847" stroke-width="1.5"/><line x1="105" y1="61" x2="105" y2="57" stroke="#6e5847" stroke-width="1.5"/>
<line x1="125" y1="61" x2="125" y2="57" stroke="#6e5847" stroke-width="1.5"/><line x1="145" y1="61" x2="145" y2="57" stroke="#6e5847" stroke-width="1.5"/>
<circle cx="50" cy="39" r="6" fill="#c9a96e" opacity="0.9"/><circle cx="75" cy="32" r="8" fill="#b5994e" opacity="0.85"/>
<circle cx="100" cy="36" r="7" fill="#d4b87a" opacity="0.9"/><circle cx="128" cy="34" r="9" fill="#b08840" opacity="0.85"/>
<circle cx="150" cy="38" r="5" fill="#c9a96e" opacity="0.9"/>
<circle cx="60" cy="73" r="2.5" fill="#c9a96e" opacity="0.7"/><circle cx="80" cy="80" r="2" fill="#d4b87a" opacity="0.6"/>
<circle cx="100" cy="76" r="2.5" fill="#b5994e" opacity="0.7"/><circle cx="120" cy="82" r="2" fill="#c9a96e" opacity="0.6"/>
<circle cx="140" cy="78" r="2.2" fill="#d4b87a" opacity="0.6"/><circle cx="90" cy="92" r="1.5" fill="#b5994e" opacity="0.4"/>
</svg>"""

# ── PWA : manifest + meta tags pour installation mobile ──
PWA_MANIFEST = {
    "name": "Classification GTR2023",
    "short_name": "GTR2023",
    "description": "Classification des sols selon le GTR 2023 - Remblai, couche de forme, compactage",
    "start_url": "/",
    "display": "standalone",
    "background_color": "#f6f1e7",
    "theme_color": "#b5651d",
    "orientation": "portrait",
    "icons": [
        {"src": "data:image/svg+xml," + LOGO_SVG.replace('"','%22').replace('#','%23').replace('\n',''), "sizes": "any", "type": "image/svg+xml"}
    ]
}

import urllib.parse
manifest_json = json.dumps(PWA_MANIFEST)
manifest_b64 = base64.b64encode(manifest_json.encode()).decode()

st.markdown(f"""
<link rel="manifest" href="data:application/json;base64,{manifest_b64}">
<meta name="mobile-web-app-capable" content="yes">
<meta name="apple-mobile-web-app-capable" content="yes">
<meta name="apple-mobile-web-app-status-bar-style" content="black-translucent">
<meta name="apple-mobile-web-app-title" content="GTR2023">
<meta name="theme-color" content="#b5651d">
<meta name="viewport" content="width=device-width, initial-scale=1.0, maximum-scale=1.0, user-scalable=no">
<style>
:root{{--ocre:#b5651d;--ocre-dark:#8a4a14;--terre:#6e5847;--fond:#f6f1e7;--carte:#fffdf8;--ligne:#d8cdb8;--texte:#2e2620;--texte-doux:#6b5f50;--accent2:#2d5d6b;}}
.stApp{{background-color:var(--fond);}}h1{{color:var(--texte)!important;text-align:center;}}h2,h3{{color:var(--accent2)!important;}}
.result-box{{background:var(--carte);border:2px solid var(--ocre);border-radius:12px;padding:20px;margin:14px 0;}}
.class-symbol{{font-size:48px;font-weight:800;color:var(--ocre);}}
.badge{{display:inline-block;font-size:12px;font-weight:700;padding:4px 10px;border-radius:999px;background:#efe6d4;color:var(--terre);border:1px solid var(--ligne);margin:2px 3px;}}
.badge-main{{background:var(--ocre)!important;color:#fff!important;border:none!important;}}
.note-box{{background:#efe6d4;border-radius:8px;padding:12px 14px;font-size:12px;color:var(--texte-doux);margin-top:10px;}}
.code-box{{background:#2e2620;color:#f6f1e7;border-radius:8px;padding:10px 14px;font-family:monospace;font-size:14px;margin:6px 0;letter-spacing:0.08em;}}
.err-field{{color:#cc3333;font-size:12px;font-weight:600;}}
.footer-dev{{background:linear-gradient(135deg,#2d5d6b,#1a3a42);color:#d4cfc5;border-radius:10px;padding:18px 20px;margin-top:20px;font-size:12px;line-height:1.6;}}
.footer-dev strong{{color:#f6f1e7;}}
.install-banner{{background:linear-gradient(135deg,var(--ocre),var(--ocre-dark));color:white;border-radius:10px;padding:14px 18px;margin:10px 0;text-align:center;font-size:13px;}}
.install-banner a{{color:#ffe0b0;font-weight:700;}}
</style>
""", unsafe_allow_html=True)

def show_header():
    st.markdown(LOGO_SVG, unsafe_allow_html=True)

CREDIT = """<div class="footer-dev">
<strong>Classification GTR2023</strong> — Développé par <strong>Serigne Mouhamadane SY</strong>,
ingénieur géotechnicien, Université Iba Der Thiam de Thiès.<br><br>
Cet outil a été conçu pour accompagner les étudiants et les professionnels du génie civil
dans la classification rapide des sols selon le GTR 2023. Il permet un gain de temps significatif
dans la phase d'étude, mais <strong>ne dispense en aucun cas de la vérification par un ingénieur
qualifié</strong>. Les résultats doivent être confrontés aux données de terrain, aux essais de laboratoire
et aux prescriptions du Guide des Terrassements (IDRRIM/Cerema, Fascicules 1 et 2, éd. 2024).
<br><br>
📩 Pour toute remarque, suggestion ou contribution :<br>
📧 <a href="mailto:smouhamadane.sy@univ-thies.sn" style="color:#ffe0b0;">smouhamadane.sy@univ-thies.sn</a><br>
💬 <a href="https://wa.me/221773367568" style="color:#ffe0b0;">WhatsApp : +221 77 336 75 68</a>
</div>"""

# ══════════════════════════════════════════════════════════════════
# DONNÉES GTR
# ══════════════════════════════════════════════════════════════════
DESC={"F1":"Limons peu plastiques, loess, silts","F2":"Sables fins argileux, limons, argiles peu plastiques",
"F3":"Argiles, argiles marneuses, limons très plastiques","F4":"Argiles très plastiques","F4+":"Argiles IP>55",
"I1":"Sables et graves très silteux","I2":"Sables et graves argileux",
"S1":"Sables propres Cu≥6","S2":"Sables propres Cu<6","S3":"Sables limoneux Cu≥6","S4":"Sables limoneux Cu<6",
"G1":"Graves peu fines Cu≥6","G2":"Graves homométriques Cu<6","G3":"Graves silteuses Cu≥6","G4":"Graves silteuses Cu<6"}
CARACTERES = {
    "F1": "Ces sols changent brutalement de consistance pour de faibles variations de teneur en eau, en particulier lorsque leur wn est proche de wOPN. Le temps de réaction aux variations de l'environnement hydrique et climatique est relativement court. Il est souvent préférable de les identifier par la VBS (imprécision de l'IP).",
    "F2": "Le caractère moyen de ces sols fait qu'ils se prêtent à l'emploi de la plus large gamme d'outils de terrassement (si la teneur en eau n'est pas trop élevée). Dès que l'IP atteint des valeurs >= 12, il constitue le critère d'identification le mieux adapté.",
    "F3": "Ces sols sont très cohérents à teneur en eau moyenne et faible, et collants ou glissants à l'état humide, d'où la difficulté de mise en oeuvre sur chantier. Leur perméabilité très réduite rend leurs variations de teneur en eau très lentes. Une augmentation de teneur en eau assez importante est nécessaire pour changer notablement leur consistance.",
    "F4": "Ces sols sont très cohérents et presque imperméables : s'ils changent de teneur en eau, c'est extrêmement lentement et avec d'importants retraits ou gonflements. Leur emploi en remblai de faible hauteur est envisageable avec un traitement adapté.",
    "F4+": "L'utilisation des argiles très plastiques en l'état n'est pas envisageable hors étude spécifique et restera limitée à des remblais de faible hauteur.",
    "I1": "La proportion de fines et la faible plasticité de ces dernières rapprochent beaucoup le comportement de ces sols de celui des sols F. Il y a lieu de préférer le critère VBS au critère IP pour l'identification.",
    "I2": "L'influence des fines est prépondérante ; le comportement du sol se rapproche de celui du sol fin ayant la même plasticité avec toutefois une plus grande sensibilité à l'eau due à la présence de la fraction sableuse en plus grande quantité.",
    "S1": "Sables propres à granulométrie étalée (Cu >= 6). Sans cohésion et perméables, généralement insensibles à l'eau mais cette insensibilité devra être confirmée (VBS, CBRi). Emploi en couche de forme non traitée : mesure de la friabilité FS nécessaire.",
    "S2": "Sables propres à granulométrie uniforme (Cu < 6). Sans cohésion et perméables, généralement insensibles à l'eau. Granulométrie mal graduée et de petit calibre, très érodables et traficabilité difficile.",
    "S3": "Sables limoneux/argileux à granulométrie étalée. La plasticité et/ou la quantité de fines rendent ces sols généralement sensibles à l'eau. Temps de réaction aux variations hydriques court. Quand sensibles à l'eau en état h ou th, difficiles à améliorer par essorage.",
    "S4": "Sables limoneux/argileux à granulométrie uniforme. Sensibles à l'eau. Granulométrie uniforme et de petit calibre, très érodables et traficabilité difficile.",
    "G1": "Graves à granulométrie étalée contenant peu de fines. Généralement insensibles à l'eau (à confirmer par VBS). Emploi en couche de forme non traitée : mesure de LA et MDE nécessaire.",
    "G2": "Graves homométriques contenant peu de fines. Sans cohésion et perméables, généralement insensibles à l'eau. Granulométrie homométrique. Mauvaise traficabilité surtout si roulés.",
    "G3": "Graves silteuses/argileuses à granulométrie étalée. Sensibilité à l'eau dépendant de la plasticité des fines. Perméables. Réaction assez rapide aux variations hydriques. Après compactage, d'autant moins érodables qu'ils sont bien gradués.",
    "G4": "Graves silteuses/argileuses mal graduées. Teneur en fines les rend impropres au drainage. Emploi en couche de forme non traitée : mesure de LA et MDE nécessaire.",
}
ETQ={"th":"très humide","h":"humide","m":"moyen","s":"sec","ts":"très sec","ins":"insensible à l'eau"}

OBS_REMBLAI = {
    "F1": {"th": "Sol non traficable. Portance quasi nulle.", "h": "Portance faible. Mise en oeuvre difficile. Matelassage a eviter en arase.", "m": "Conditions optimales de mise en oeuvre.", "s": "IPI eleve mais risque de desordres en presence d'eau d'infiltration (remblais moyens/grands).", "ts": "Non reutilisable en conditions courantes."},
    "F2": {"th": "Sol non traficable.", "h": "Sols difficiles a mettre en oeuvre. Depot provisoire et drainage non envisageables en climat francais moyen.", "m": "Conditions optimales.", "s": "Surcompactage possible mais attention aux remblais de grande hauteur.", "ts": "Non reutilisable en conditions courantes."},
    "F3": {"th": "Non reutilisable.", "h": "Tres coherents et collants. L'aeration est peu efficace du fait de la faible permeabilite.", "m": "Conditions optimales mais sols tres coherents.", "s": "Humidification necessaire, penetration de l'eau tres lente.", "ts": "Non reutilisable."},
    "F4": {"th": "Non reutilisable.", "h": "Emploi en remblai de faible hauteur uniquement, avec traitement.", "m": "Traitement necessaire meme en etat moyen pour remblai de faible hauteur.", "s": "Humidification et traitement necessaires.", "ts": "Non reutilisable."},
    "F4+": {"th": "Non reutilisable.", "h": "Non reutilisable sauf etude specifique.", "m": "Etude specifique obligatoire, remblai de faible hauteur uniquement.", "s": "Etude specifique obligatoire.", "ts": "Non reutilisable."},
    "I1": {"th": "Non traficable.", "h": "Comportement proche des sols F. Mise en oeuvre difficile.", "m": "Conditions optimales.", "s": "Humidification ou surcompactage.", "ts": "Non reutilisable."},
    "I2": {"th": "Non traficable.", "h": "Influence des fines preponderante. Sensibilite a l'eau elevee.", "m": "Conditions optimales.", "s": "Humidification necessaire.", "ts": "Non reutilisable."},
    "S1": {"ins": "Insensible a l'eau. Mise en oeuvre sans difficulte particuliere.", "th": "Difficile a compacter si sature.", "h": "Arrosage a eviter, risque de liquefaction.", "m": "Conditions optimales.", "s": "Arrosage leger pour eviter la segregation.", "ts": "Humidification necessaire."},
    "S2": {"ins": "Insensible a l'eau. Traficabilite difficile du fait de la granulometrie uniforme.", "th": "Tres difficile a compacter.", "h": "Mise en oeuvre delicate.", "m": "Conditions optimales mais traficabilite mediocre.", "s": "Arrosage.", "ts": "Humidification necessaire."},
    "S3": {"ins": "Insensible a l'eau. Mise en oeuvre courante.", "th": "Non traficable.", "h": "Essorage difficile. Traitement possible.", "m": "Conditions optimales.", "s": "Arrosage pour maintien.", "ts": "Humidification necessaire."},
    "S4": {"ins": "Insensible a l'eau malgre granulometrie uniforme.", "th": "Non traficable.", "h": "Essorage difficile, traficabilite mediocre.", "m": "Conditions optimales mais erodable.", "s": "Arrosage.", "ts": "Humidification necessaire."},
    "G1": {"ins": "Insensible a l'eau. Materiau de bonne qualite pour remblai et couche de forme.", "th": "Rare. Drainage suffisant.", "h": "Mise en oeuvre possible.", "m": "Conditions optimales. Excellent materiau de remblai.", "s": "Arrosage leger.", "ts": "Humidification."},
    "G2": {"ins": "Insensible a l'eau. Traficabilite mediocre si roules.", "th": "Rare.", "h": "Mise en oeuvre possible.", "m": "Conditions optimales mais traficabilite limitee.", "s": "Arrosage.", "ts": "Humidification."},
    "G3": {"ins": "Insensible a l'eau. Bon materiau.", "th": "Non traficable.", "h": "Sensibilite a l'eau selon plasticite des fines. Aeration possible.", "m": "Conditions optimales.", "s": "Arrosage pour maintien.", "ts": "Humidification necessaire."},
    "G4": {"ins": "Insensible a l'eau malgre granulometrie uniforme.", "th": "Non traficable.", "h": "Mise en oeuvre delicate.", "m": "Conditions optimales.", "s": "Arrosage.", "ts": "Humidification necessaire."},
}

# PST simplified determination (Fascicule 1, Chapitre 4)
def get_pst(sc, etat, famille):
    """Simplified PST determination based on GTR 2023."""
    if famille == "R" or sc.startswith(("CH","Li","Cl","Sa","Co","SR","Vo","Me")):
        return "PST0", "Materiau rocheux - PST fonction de la fragmentation et de l'etat."
    if sc in ("S1","S2","G1","G2"):
        if etat == "ins": return "PST1", "Sol grenus insensibles a l'eau."
        if etat == "m": return "PST1", "Sol grenus etat moyen."
        if etat in ("h","s"): return "PST2", "Sol grenus etat h ou s."
        return "PST3", "Sol grenus etat defavorable."
    if sc in ("S3","S4","G3","G4"):
        if etat == "ins": return "PST1", "Sol grenus insensibles a l'eau."
        if etat == "m": return "PST2", "Sol grenus sensibles a l'eau, etat moyen."
        if etat == "h": return "PST3", "Sol grenus sensibles a l'eau, etat humide."
        if etat == "s": return "PST2", "Sol grenus sensibles, etat sec - humidification necessaire."
        return "PST4", "Sol grenus etat defavorable."
    if famille in ("F","I"):
        if etat == "m": return "PST3", "Sol fin/intermediaire etat moyen."
        if etat == "h": return "PST4", "Sol fin/intermediaire etat humide."
        if etat == "s": return "PST3", "Sol fin/intermediaire etat sec - humidification necessaire."
        return "PST5", "Sol fin/intermediaire etat defavorable - etude specifique."
    return "PST3", "Classification PST par defaut - a preciser par etude."
SEUILS_WN={"F1":(1.25,1.10,0.90,0.70),"F2":(1.30,1.10,0.90,0.70),"F3":(1.40,1.20,0.90,0.70),
"F4":(1.40,1.20,0.90,0.70),"I1":(1.25,1.10,0.90,0.60),"I2":(1.30,1.10,0.90,0.70)}
SEUILS_IPI={"F1":(3,8,25),"F2":(2,6,15),"F3":(2,4,10),"F4":(1,3,10),"I1":(5,12,30),"I2":(4,10,25)}
SEUILS_IC={"F2":(0.95,1.05,1.15,1.30),"F3":(0.85,1.00,1.10,1.25),"F4":(0.80,1.00,1.10,1.20),"I2":(0.85,1.00,1.15,1.25)}
CODES={"E":{0:"Pas de condition particulière",1:"Extraction en couches (0,1-0,3 m)",2:"Extraction frontale"},
"G":{0:"Pas de condition particulière",1:"Élimination Lmax>800mm",2:"Élimination Lmax>250mm",3:"Fragmentation complémentaire"},
"W":{0:"Pas de condition particulière",1:"Réduction wn par aération",2:"Essorage dépôt provisoire",3:"Arrosage maintien",4:"Humidification"},
"T":{0:"Pas de condition particulière",1:"Traitement réactif adapté",2:"Traitement chaux seule"},
"R":{0:"Pas de condition particulière",1:"Couches minces (20-30cm)",2:"Couches moyennes (30-50cm)"},
"C":{1:"Compactage intense",2:"Compactage moyen",3:"Compactage faible"},
"H":{0:"Pas de limitation",1:"H≤5m",2:"H≤10m"}}
ROCHES={"CH":"Craie","Li":"Calcaires","Cl":"Roches argileuses","Sa":"Grès","Co":"Conglomérats",
"SR":"Roches salines","Vo":"Roches magmatiques","Me":"Roches métamorphiques"}

COMPACTAGE_QS = {
    "F1": {
        "q4": {"V3": {"qs":0.085,"e":0.30,"v":2.5}, "V4": {"qs":0.100,"e":0.40,"v":2.5}, "V5": {"qs":0.130,"e":0.45,"v":2.5},
               "VP3": {"qs":0.085,"e":0.30,"v":2.5}, "VP4": {"qs":0.100,"e":0.30,"v":3.5}, "VP5": {"qs":0.130,"e":0.30,"v":4.0},
               "P2": {"qs":0.065,"e":0.35,"v":5.0}, "P3": {"qs":0.095,"e":0.45,"v":5.0},
               "SP1": {"qs":0.040,"e":0.20,"v":8.0}, "SP2": {"qs":0.070,"e":0.30,"v":8.0}},
        "q3": {"V3": {"qs":0.050,"e":0.30,"v":2.0}, "V4": {"qs":0.065,"e":0.30,"v":2.0}, "V5": {"qs":0.085,"e":0.30,"v":2.0},
               "VP3": {"qs":0.050,"e":0.20,"v":2.0}, "VP4": {"qs":0.065,"e":0.30,"v":2.5}, "VP5": {"qs":0.085,"e":0.30,"v":3.0},
               "P2": {"qs":0.050,"e":0.20,"v":5.0}, "P3": {"qs":0.035,"e":0.30,"v":5.0},
               "SP1": {"qs":0.035,"e":0.25,"v":8.0}},
    },
    "F2": {
        "q4": {"V3": {"qs":0.060,"e":0.30,"v":2.0}, "V4": {"qs":0.080,"e":0.35,"v":2.5}, "V5": {"qs":0.100,"e":0.40,"v":2.5},
               "VP3": {"qs":0.060,"e":0.25,"v":2.5}, "VP4": {"qs":0.080,"e":0.30,"v":3.0}, "VP5": {"qs":0.100,"e":0.30,"v":3.5},
               "P2": {"qs":0.050,"e":0.25,"v":5.0}, "P3": {"qs":0.080,"e":0.35,"v":5.0},
               "SP1": {"qs":0.040,"e":0.20,"v":8.0}, "SP2": {"qs":0.065,"e":0.30,"v":8.0}},
        "q3": {"V4": {"qs":0.050,"e":0.25,"v":2.0}, "V5": {"qs":0.065,"e":0.30,"v":2.0},
               "VP4": {"qs":0.050,"e":0.25,"v":2.5}, "VP5": {"qs":0.065,"e":0.30,"v":3.0},
               "P3": {"qs":0.050,"e":0.30,"v":5.0}},
    },
    "F3": {
        "q4": {"V3": {"qs":0.055,"e":0.25,"v":2.0}, "V4": {"qs":0.070,"e":0.30,"v":2.5}, "V5": {"qs":0.085,"e":0.35,"v":2.5},
               "VP3": {"qs":0.070,"e":0.30,"v":3.5}, "VP4": {"qs":0.085,"e":0.30,"v":4.0}, "VP5": {"qs":0.110,"e":0.30,"v":5.0},
               "P2": {"qs":0.040,"e":0.25,"v":5.0}, "P3": {"qs":0.060,"e":0.35,"v":5.0},
               "SP1": {"qs":0.040,"e":0.20,"v":8.0}, "SP2": {"qs":0.070,"e":0.30,"v":8.0}},
        "q3": {"V4": {"qs":0.045,"e":0.25,"v":2.0}, "V5": {"qs":0.055,"e":0.30,"v":2.0},
               "VP4": {"qs":0.055,"e":0.25,"v":3.0}, "VP5": {"qs":0.070,"e":0.30,"v":3.5},
               "P3": {"qs":0.040,"e":0.25,"v":5.0}},
    },
    "F4": {
        "q4": {"V3": {"qs":0.040,"e":0.20,"v":2.0}, "V4": {"qs":0.055,"e":0.25,"v":2.0}, "V5": {"qs":0.070,"e":0.30,"v":2.0},
               "VP3": {"qs":0.055,"e":0.25,"v":3.0}, "VP4": {"qs":0.070,"e":0.25,"v":3.5}, "VP5": {"qs":0.085,"e":0.30,"v":4.0},
               "P2": {"qs":0.030,"e":0.20,"v":5.0}, "P3": {"qs":0.050,"e":0.30,"v":5.0},
               "SP1": {"qs":0.035,"e":0.20,"v":8.0}, "SP2": {"qs":0.060,"e":0.25,"v":8.0}},
        "q3": {"V4": {"qs":0.035,"e":0.20,"v":2.0}, "V5": {"qs":0.045,"e":0.25,"v":2.0},
               "VP4": {"qs":0.045,"e":0.20,"v":3.0}, "VP5": {"qs":0.055,"e":0.25,"v":3.5},
               "P3": {"qs":0.035,"e":0.25,"v":5.0}},
    },
    "I1": {
        "q4": {"V3": {"qs":0.090,"e":0.30,"v":2.5}, "V4": {"qs":0.110,"e":0.40,"v":2.5}, "V5": {"qs":0.140,"e":0.45,"v":2.5},
               "VP3": {"qs":0.090,"e":0.30,"v":3.0}, "VP4": {"qs":0.110,"e":0.35,"v":3.5}, "VP5": {"qs":0.140,"e":0.35,"v":4.0},
               "P2": {"qs":0.070,"e":0.35,"v":5.0}, "P3": {"qs":0.100,"e":0.45,"v":5.0}},
        "q3": {"V4": {"qs":0.070,"e":0.30,"v":2.0}, "V5": {"qs":0.090,"e":0.35,"v":2.0},
               "VP4": {"qs":0.070,"e":0.30,"v":3.0}, "VP5": {"qs":0.090,"e":0.30,"v":3.5},
               "P3": {"qs":0.065,"e":0.35,"v":5.0}},
    },
    "I2": {
        "q4": {"V3": {"qs":0.065,"e":0.30,"v":2.0}, "V4": {"qs":0.085,"e":0.35,"v":2.5}, "V5": {"qs":0.105,"e":0.40,"v":2.5},
               "VP3": {"qs":0.065,"e":0.25,"v":2.5}, "VP4": {"qs":0.085,"e":0.30,"v":3.0}, "VP5": {"qs":0.105,"e":0.30,"v":3.5},
               "P2": {"qs":0.055,"e":0.25,"v":5.0}, "P3": {"qs":0.085,"e":0.35,"v":5.0},
               "SP1": {"qs":0.040,"e":0.20,"v":8.0}, "SP2": {"qs":0.065,"e":0.30,"v":8.0}},
        "q3": {"V4": {"qs":0.055,"e":0.25,"v":2.0}, "V5": {"qs":0.070,"e":0.30,"v":2.0},
               "VP4": {"qs":0.055,"e":0.25,"v":2.5}, "VP5": {"qs":0.070,"e":0.30,"v":3.0},
               "P3": {"qs":0.055,"e":0.30,"v":5.0}},
    },
    "S": {
        "q4": {"V3": {"qs":0.125,"e":0.35,"v":3.0}, "V4": {"qs":0.150,"e":0.45,"v":3.0}, "V5": {"qs":0.185,"e":0.50,"v":3.0},
               "P2": {"qs":0.100,"e":0.40,"v":5.0}, "P3": {"qs":0.150,"e":0.50,"v":5.0},
               "PQ3": {"qs":0.040,"e":0.20,"v":1.0}, "PQ4": {"qs":0.060,"e":0.25,"v":1.0}},
        "q3": {"V4": {"qs":0.100,"e":0.35,"v":2.5}, "V5": {"qs":0.125,"e":0.40,"v":2.5},
               "P3": {"qs":0.100,"e":0.40,"v":5.0}},
    },
    "G": {
        "q4": {"V3": {"qs":0.145,"e":0.40,"v":3.0}, "V4": {"qs":0.175,"e":0.50,"v":3.0}, "V5": {"qs":0.215,"e":0.55,"v":3.0},
               "P2": {"qs":0.120,"e":0.45,"v":5.0}, "P3": {"qs":0.180,"e":0.60,"v":5.0},
               "PQ3": {"qs":0.050,"e":0.25,"v":1.0}, "PQ4": {"qs":0.070,"e":0.30,"v":1.0}},
        "q3": {"V4": {"qs":0.115,"e":0.40,"v":2.5}, "V5": {"qs":0.145,"e":0.45,"v":2.5},
               "P3": {"qs":0.120,"e":0.45,"v":5.0}},
    },
}

# ══════════════════════════════════════════════════════════════════
# FONCTIONS
# ══════════════════════════════════════════════════════════════════
def classify(p63,p2,ip,vbs,cu):
    if p63>35:
        if ip and ip>55: return "F","F4+"
        if (ip and ip>40)or(vbs and vbs>8): return "F","F4"
        if (ip and ip>22)or(vbs and 6<vbs<=8): return "F","F3"
        if (ip and ip>12)or(vbs and 2.5<vbs<=6): return "F","F2"
        return "F","F1"
    elif p63>=15:
        return ("I","I2") if ((vbs and vbs>1.5)or(ip and ip>12)) else ("I","I1")
    else:
        if p2 is None:
            return ("S","S3") if (cu and cu>=6) else ("S","S4")
        fs=max(p2-p63,0); fg=max(100-p2,0)
        if fs>=fg:
            if p63<=5: return ("S","S1") if(cu and cu>=6) else ("S","S2")
            return ("S","S3") if(cu and cu>=6) else ("S","S4")
        else:
            if p63<=5: return ("G","G1") if(cu and cu>=6) else ("G","G2")
            return ("G","G3") if(cu and cu>=6) else ("G","G4")

def calc_etat(sc,wn,wopn,ipi,ic,p2):
    res=[]
    if wn is None and wopn is None and ipi is None and ic is None:
        return res
    if wn is not None and wopn and wopn>0:
        r=wn/wopn
        if sc in SEUILS_WN: th,h,m,s=SEUILS_WN[sc]
        elif sc[0] in "SG": th,h,m,s=(1.25,1.10,0.90,0.50) if(p2 and p2>70) else (1.25,1.10,0.90,0.60)
        else: th,h,m,s=(1.25,1.10,0.90,0.70)
        e="th" if r>=th else "h" if r>=h else "m" if r>=m else "s" if r>=s else "ts"
        res.append(("wn/wOPN",e,f"{r:.2f}"))
    if ipi is not None:
        if sc in SEUILS_IPI: tm,hm,mm=SEUILS_IPI[sc]
        elif sc[0]=="S": tm,hm,mm=(4,8,None) if(p2 and p2>70) else (6,12,None)
        elif sc[0]=="G": tm,hm,mm=(7,15,30)
        else: tm,hm,mm=(3,8,25)
        e="th" if ipi<=tm else "h" if ipi<=hm else "m"
        res.append(("IPI",e,str(ipi)))
    if ic is not None and sc in SEUILS_IC:
        a,b,c,d=SEUILS_IC[sc]
        e="th" if ic<=a else "h" if ic<=b else "m" if ic<=c else "s" if ic<=d else "ts"
        res.append(("Ic",e,f"{ic:.2f}"))
    return res

def resolve_etat(etats):
    """Resolve divergent hydric state methods. Returns (etat, divergence_message_or_None)."""
    if not etats:
        return "m", None
    if len(etats) == 1:
        return etats[0][1], None
    states = {e[1] for e in etats}
    if len(states) == 1:
        return etats[0][1], None
    # Divergence: apply priority rules
    ipi_etat = None; wn_etat = None; ic_etat = None
    for meth, et, val in etats:
        if meth == "IPI": ipi_etat = et
        elif meth == "wn/wOPN": wn_etat = et
        elif meth == "Ic": ic_etat = et
    # IPI more reliable for wet states
    if ipi_etat in ("h", "th"):
        msg = f"Divergence etats hydriques. IPI privilegie (etat humide/tres humide plus fiable par IPI)."
        return ipi_etat, msg
    # wn/wOPN more reliable for dry states
    if wn_etat in ("s", "ts"):
        msg = f"Divergence etats hydriques. wn/wOPN privilegie (etat sec/tres sec plus fiable par wn/wOPN)."
        return wn_etat, msg
    # Otherwise take first
    methods_str = ", ".join(f"{m}={v}->{e}" for m, e, v in etats)
    return etats[0][1], f"Divergence etats hydriques ({methods_str}). Premier resultat retenu."

def check_ins(sc,p63,vbs,cbri):
    if sc[0] not in "SG": return False
    if p63<=5 and vbs is not None and vbs<0.2: return True
    if 5<p63<=10 and vbs is not None and vbs<0.1: return True
    if 5<p63<=10 and vbs is not None and 0.1<=vbs<0.2 and cbri and cbri>20: return True
    if sc[0]=="G" and 10<p63<=12 and vbs is not None and vbs<0.1 and cbri and cbri>20: return True
    return False

def get_remblai(sc,etat,fam):
    if etat in("th","ts"): return f"NON — État {ETQ[etat]}. Étude spécifique nécessaire."
    if sc=="F4+" and etat=="h": return "NON — Étude spécifique obligatoire."
    if etat=="h":
        d={"Pluie forte (++)":"NON","Pluie faible (+)":"NON"}
        if fam in("F","I"):
            d["Ni pluie ni évap. (=)"]=[("Traitement chaux",{"E":0,"G":0,"W":0,"T":2,"R":0,"C":2,"H":0}),("En l'état",{"E":0,"G":0,"W":0,"T":0,"R":0,"C":3,"H":1})]
            d["Évaporation (-)"]=[("Aération",{"E":1,"G":0,"W":1,"T":0,"R":1,"C":2,"H":2}),("Traitement chaux",{"E":0,"G":0,"W":0,"T":2,"R":0,"C":2,"H":0})]
        else:
            d["Ni pluie ni évap. (=)"]=[("En l'état",{"E":0,"G":0,"W":0,"T":0,"R":0,"C":3,"H":1})]
            d["Évaporation (-)"]=[("Aération",{"E":1,"G":0,"W":1,"T":0,"R":1,"C":2,"H":2})]
        return d
    if etat in("m","ins"):
        return {"Pluie forte (++)":[("En l'état",{"E":0,"G":0,"W":0,"T":0,"R":0,"C":2,"H":0})],"Pluie faible (+)":[("En l'état",{"E":0,"G":0,"W":0,"T":0,"R":0,"C":2,"H":0})],"Ni pluie ni évap. (=)":[("En l'état",{"E":0,"G":0,"W":0,"T":0,"R":0,"C":1,"H":0})],"Évaporation (-)":[("Arrosage",{"E":0,"G":0,"W":3,"T":0,"R":0,"C":1,"H":0})]}
    if etat=="s":
        return {"Pluie forte (++)":[("En l'état",{"E":0,"G":0,"W":0,"T":0,"R":0,"C":1,"H":0})],"Pluie faible (+)":[("En l'état",{"E":0,"G":0,"W":0,"T":0,"R":0,"C":1,"H":0})],"Ni pluie ni évap. (=)":[("Humidification",{"E":0,"G":0,"W":4,"T":0,"R":0,"C":1,"H":2})],"Évaporation (-)":[("Humidification",{"E":0,"G":0,"W":4,"T":0,"R":0,"C":1,"H":1})]}
    return "Se reporter au Fascicule 2."

def get_compacteurs(fam,etat):
    if fam in("F","I"):
        q4=[("VP3-VP5","Vibrant pieds dameurs"),("SP1-SP2","Statique pieds dameurs"),("V3-V5","Vibrant cylindre lisse"),("P2-P3","Pneus")]
        q3=[("VP4-VP5","Vibrant pieds dameurs"),("V4-V5","Vibrant cylindre lisse"),("P3","Pneus CR>60kN")]
        if etat=="h": q4=[("P2-P3","Pneus sols humides"),("VP3-VP5","Vibrant pieds dameurs")]
    elif fam=="S":
        q4=[("V3-V5","Vibrant cylindre lisse"),("P2-P3","Pneus"),("PQ3-PQ4","Plaques vibrantes")]
        q3=[("V4-V5","Vibrant cylindre lisse"),("P3","Pneus")]
    else:
        q4=[("V3-V5","Vibrant cylindre lisse"),("P2-P3","Pneus")]
        q3=[("V4-V5","Vibrant cylindre lisse"),("P3","Pneus")]
    return {"q4":q4,"q3":q3}

def show_synoptique(famille, sc, p63):
    """Visual synoptique showing where the soil falls in GTR classification"""
    st.markdown("**Position dans la classification GTR 2023 :**")
    if p63 is not None:
        if p63 > 35:
            level1 = f"Passant 63 µm = {p63:.1f}% > 35% → **Sols fins (F)**"
        elif p63 >= 15:
            level1 = f"Passant 63 µm = {p63:.1f}% (15-35%) → **Sols intermédiaires (I)**"
        else:
            level1 = f"Passant 63 µm = {p63:.1f}% < 15% → **Sols sableux/graveleux (S/G)**"
        st.markdown(f"1️⃣ {level1}")
    st.markdown(f"2️⃣ Sous-classe → **{sc}** ({DESC.get(sc, sc)})")
    classes_in_family = {
        "F": ["F1","F2","F3","F4","F4+"],
        "I": ["I1","I2"],
        "S": ["S1","S2","S3","S4"],
        "G": ["G1","G2","G3","G4"],
    }
    if famille in classes_in_family:
        cols = st.columns(len(classes_in_family[famille]))
        for i, cls in enumerate(classes_in_family[famille]):
            with cols[i]:
                if cls == sc:
                    st.markdown(f'<div style="background:#b5651d;color:white;text-align:center;padding:8px;border-radius:6px;font-weight:700;font-size:16px;">{cls}</div>', unsafe_allow_html=True)
                else:
                    st.markdown(f'<div style="background:#efe6d4;color:#6e5847;text-align:center;padding:8px;border-radius:6px;font-size:14px;">{cls}</div>', unsafe_allow_html=True)

def get_qs_table(sc, famille):
    """Get Q/S compaction table for the given soil class or family."""
    return COMPACTAGE_QS.get(sc) or COMPACTAGE_QS.get(famille)

def sanitize(text):
    """Remplace les caractères Unicode non supportés par Helvetica (Latin-1)."""
    if not isinstance(text, str): return str(text)
    replacements = {
        "—": "-", "–": "-", "‘": "'", "’": "'",
        "“": '"', "”": '"', "…": "...",
        "≥": ">=", "≤": "<=", "≠": "!=",
        "ρ": "rho", "µ": "u", "→": "->",
        "✅": "OK", "❌": "NON", "⚠️": "!",
        "⚠": "!", "•": "-",
    }
    for old, new in replacements.items():
        text = text.replace(old, new)
    return text

def build_excel(projet_info, sondages):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = Workbook()
    ws = wb.active
    ws.title = "Synthese"
    headers = ["Sondage", "Classe", "Etat", "Remblai", "CdF", "Compactage", "PST"]
    header_fill = PatternFill(start_color="B5651D", end_color="B5651D", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = header_fill; c.font = header_font; c.alignment = Alignment(horizontal="center")
    for row, s in enumerate(sondages, 2):
        ws.cell(row=row, column=1, value=s["sondage_id"])
        ws.cell(row=row, column=2, value=s["symbole"])
        ws.cell(row=row, column=3, value=ETQ.get(s.get("etat",""), ""))
        ws.cell(row=row, column=4, value="Oui" if s.get("do_remblai") and not isinstance(s.get("remblai_cond"), str) else "Etude")
        cdf = s.get("cdf_result", {})
        ws.cell(row=row, column=5, value=cdf.get("texte", "-"))
        ws.cell(row=row, column=6, value="Oui" if s.get("do_compact") else "-")
        ws.cell(row=row, column=7, value=s.get("pst", "-"))
    for col in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)
    # Sheet 2: Details
    ws2 = wb.create_sheet("Details")
    detail_headers = ["Sondage", "Type", "Famille", "Sous-classe", "Etat", "Parametres", "Observations remblai"]
    for col, h in enumerate(detail_headers, 1):
        c = ws2.cell(row=1, column=col, value=h)
        c.fill = header_fill; c.font = header_font
    for row, s in enumerate(sondages, 2):
        ws2.cell(row=row, column=1, value=s["sondage_id"])
        ws2.cell(row=row, column=2, value=s.get("type", ""))
        ws2.cell(row=row, column=3, value=s.get("famille", ""))
        ws2.cell(row=row, column=4, value=s.get("sc", ""))
        ws2.cell(row=row, column=5, value=ETQ.get(s.get("etat",""), ""))
        ws2.cell(row=row, column=6, value="; ".join(f"{k}={v}" for k, v in s.get("params", {}).items()))
        ws2.cell(row=row, column=7, value=s.get("obs_remblai", ""))
    for col in ws2.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws2.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

def draw_coupe(pdf, sondages):
    """Draw a simplified geotechnical cross-section"""
    pdf.add_page()
    pdf.chapter_title("", "COUPE GEOTECHNIQUE SCHEMATIQUE")
    sondage_groups = {}
    for s in sondages:
        sid = s["sondage_id"]
        if sid not in sondage_groups:
            sondage_groups[sid] = []
        sondage_groups[sid].append(s)
    if not sondage_groups:
        return
    margin_left = 20
    col_width = min(35, 170 / max(len(sondage_groups), 1))
    max_depth = 0
    for layers in sondage_groups.values():
        for s in layers:
            d = s.get("prof_a", 0)
            try:
                val = float(str(d).replace("m","")) if isinstance(d, str) else float(d or 0)
                max_depth = max(max_depth, val)
            except:
                max_depth = max(max_depth, 5)
    if max_depth == 0:
        max_depth = 5
    y_start = pdf.get_y() + 5
    scale = 100 / max_depth
    colors = {"F": (210,180,140), "I": (180,200,160), "S": (240,220,150), "G": (190,190,190), "R": (160,160,180), "O": (120,100,80), "VC": (170,170,170)}
    for idx, (sid, layers) in enumerate(sondage_groups.items()):
        x = margin_left + idx * (col_width + 5)
        pdf.set_font("Helvetica", "B", 7)
        pdf.set_xy(x, y_start - 5)
        pdf.cell(col_width, 4, sanitize(sid), align="C")
        y_current = y_start
        for s in layers:
            h_de = s.get("prof_de", 0) or 0
            h_a = s.get("prof_a", 3) or 3
            try:
                h_de = float(str(h_de).replace("m","")) if isinstance(h_de, str) else float(h_de)
                h_a = float(str(h_a).replace("m","")) if isinstance(h_a, str) else float(h_a)
            except:
                h_de = 0; h_a = 3
            h_px = (h_a - h_de) * scale
            if h_px < 8: h_px = 8
            fam = s.get("famille", "F")
            r, g, b = colors.get(fam, (200, 200, 200))
            pdf.set_fill_color(r, g, b)
            pdf.set_draw_color(100, 100, 100)
            pdf.rect(x, y_current, col_width, h_px, "DF")
            pdf.set_font("Helvetica", "B", 6)
            pdf.set_xy(x, y_current + 1)
            pdf.cell(col_width, 3, sanitize(s.get("symbole", "")), align="C")
            pdf.set_font("Helvetica", "", 5)
            pdf.set_xy(x, y_current + 4)
            pdf.cell(col_width, 3, sanitize(f"{h_de}-{h_a}m"), align="C")
            y_current += h_px
    pdf.set_font("Helvetica", "", 6)
    pdf.set_draw_color(0, 0, 0)
    for d in range(int(max_depth) + 1):
        y = y_start + d * scale
        pdf.line(margin_left - 5, y, margin_left - 2, y)
        pdf.set_xy(margin_left - 18, y - 2)
        pdf.cell(12, 4, f"{d}m", align="R")
    pdf.ln(max_depth * scale + 15)
    pdf.set_font("Helvetica", "B", 8)
    pdf.cell(0, 5, sanitize("Legende :"), new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 7)
    for fam, (r, g, b) in colors.items():
        pdf.set_fill_color(r, g, b)
        pdf.rect(pdf.get_x(), pdf.get_y(), 8, 4, "F")
        pdf.set_xy(pdf.get_x() + 10, pdf.get_y())
        full_name = {"F":"Sols fins","I":"Intermediaires","S":"Sableux","G":"Graveleux","R":"Rocheux","O":"Organiques","VC":"Gros elements"}.get(fam, fam)
        pdf.cell(30, 4, sanitize(f"{fam} - {full_name}"))
        pdf.ln(5)

def build_pdf(projet_info, sondages):
    from fpdf import FPDF
    class PDF(FPDF):
        def s_cell(self,w,h,txt="",**kw):
            super().cell(w,h,sanitize(txt),**kw)
        def s_multi(self,w,h,txt="",**kw):
            super().multi_cell(w,h,sanitize(txt),**kw)
        def header(self):
            self.set_font("Helvetica","B",9); self.s_cell(0,5,"UNIVERSITE IBA DER THIAM DE THIES",align="C",new_x="LMARGIN",new_y="NEXT")
            self.set_font("Helvetica","",7); self.s_cell(0,4,f"Projet : {projet_info.get('projet','')} - {projet_info.get('site','')}",align="C",new_x="LMARGIN",new_y="NEXT")
            self.line(10,20,200,20); self.ln(4)
        def footer(self):
            self.set_y(-15); self.set_font("Helvetica","I",7)
            self.s_cell(0,10,f"Classification GTR2023 - S.M. SY / UIDT - Page {self.page_no()}/{{nb}} - {datetime.now().strftime('%d/%m/%Y')}",align="C")
        def chapter_title(self,num,title):
            self.set_font("Helvetica","B",13); self.set_text_color(45,93,107)
            self.s_cell(0,10,f"{num}. {title}" if num else title,new_x="LMARGIN",new_y="NEXT"); self.set_text_color(0,0,0); self.ln(2)
        def sub_title(self,t):
            self.set_font("Helvetica","B",10); self.set_text_color(138,74,20)
            self.s_cell(0,7,f"  {t}",new_x="LMARGIN",new_y="NEXT"); self.set_text_color(0,0,0)
        def param_table(self,params):
            self.set_font("Helvetica","B",8); self.set_fill_color(239,230,212)
            self.s_cell(80,6,"Parametre",border=1,fill=True); self.s_cell(50,6,"Valeur",border=1,fill=True,new_x="LMARGIN",new_y="NEXT")
            self.set_font("Helvetica","",8)
            for k,v in params.items():
                self.s_cell(80,5,k,border=1); self.s_cell(50,5,str(v),border=1,new_x="LMARGIN",new_y="NEXT")
            self.ln(2)
        def code_row(self,codes):
            self.set_font("Helvetica","B",7); self.set_fill_color(46,38,32); self.set_text_color(255,255,255)
            for r in "EGWTRC": self.s_cell(14,5,r,border=1,fill=True,align="C")
            self.s_cell(14,5,"H",border=1,fill=True,align="C",new_x="LMARGIN",new_y="NEXT")
            self.set_text_color(0,0,0); self.set_font("Courier","B",9)
            for r in ["E","G","W","T","R","C","H"]: self.s_cell(14,6,str(codes.get(r,0)),border=1,align="C")
            self.ln(); self.set_font("Helvetica","",7)
            for r,v in codes.items():
                if v!=0: self.s_cell(0,4,f"    {r}{v}: {CODES[r][v]}",new_x="LMARGIN",new_y="NEXT")
            self.ln(2)

    pdf=PDF(); pdf.alias_nb_pages(); pdf.set_auto_page_break(auto=True,margin=20)

    # PAGE DE TITRE
    pdf.add_page(); pdf.ln(60)
    pdf.set_font("Helvetica","B",26); pdf.set_text_color(138,74,20)
    pdf.s_cell(0,14,"CLASSIFICATION GTR 2023",align="C",new_x="LMARGIN",new_y="NEXT")
    pdf.set_font("Helvetica","",13); pdf.set_text_color(110,88,71)
    pdf.s_cell(0,10,"Rapport de classification et conditions d'utilisation",align="C",new_x="LMARGIN",new_y="NEXT")
    pdf.ln(8); pdf.set_text_color(0,0,0); pdf.set_font("Helvetica","",11)
    pdf.s_cell(0,7,f"Projet : {projet_info.get('projet','')}",align="C",new_x="LMARGIN",new_y="NEXT")
    pdf.s_cell(0,7,f"Site : {projet_info.get('site','')}",align="C",new_x="LMARGIN",new_y="NEXT")
    pdf.s_cell(0,7,f"Ingenieur : {projet_info.get('ingenieur','')}",align="C",new_x="LMARGIN",new_y="NEXT")
    pdf.s_cell(0,7,f"Nombre de sondages : {len(sondages)}",align="C",new_x="LMARGIN",new_y="NEXT")
    pdf.s_cell(0,7,f"Date : {datetime.now().strftime('%d/%m/%Y')}",align="C",new_x="LMARGIN",new_y="NEXT")
    pdf.ln(10); pdf.set_font("Helvetica","I",9); pdf.set_text_color(107,95,80)
    pdf.s_multi(0,4,"Referentiel : Guide des Terrassements - Ed. 2024 - IDRRIM/Cerema - Fascicules 1 et 2\nConforme NF EN 16907-2",align="C")
    pdf.ln(6); pdf.set_text_color(170,51,51); pdf.set_font("Helvetica","BI",8)
    pdf.s_cell(0,5,"Document non officiel - A valider par un geotechnicien qualifie",align="C",new_x="LMARGIN",new_y="NEXT")

    # TABLE DES MATIÈRES
    pdf.add_page(); pdf.chapter_title("","TABLE DES MATIERES"); pdf.set_font("Helvetica","",10)
    for i,s in enumerate(sondages,1):
        pdf.set_font("Helvetica","B",10)
        pdf.s_cell(0,6,f"Sondage {i} - {s['sondage_id']} : {s['symbole']}",new_x="LMARGIN",new_y="NEXT")
        pdf.set_font("Helvetica","",9)
        for sub in ["Classification","Remblai","Couche de forme","Compactage"]:
            if sub=="Remblai" and not s.get("do_remblai"): continue
            if sub=="Couche de forme" and not s.get("do_cdf"): continue
            if sub=="Compactage" and not s.get("do_compact"): continue
            pdf.s_cell(0,5,f"    {sub}",new_x="LMARGIN",new_y="NEXT")

    # CHAQUE SONDAGE
    for i,s in enumerate(sondages,1):
        pdf.add_page()
        pdf.set_font("Helvetica","B",16); pdf.set_text_color(45,93,107)
        depth_str_pdf = ""
        if s.get("prof_de") is not None and s.get("prof_a") is not None:
            depth_str_pdf = f" ({s['prof_de']}-{s['prof_a']} m)"
        pdf.s_cell(0,12,f"SONDAGE {i} - {s['sondage_id']}{depth_str_pdf}",align="C",new_x="LMARGIN",new_y="NEXT")
        pdf.set_text_color(0,0,0)

        # Classification
        pdf.chapter_title(f"{i}.1","Classification")
        pdf.set_font("Helvetica","B",18); pdf.set_text_color(138,74,20)
        pdf.s_cell(0,12,f"  {s['symbole']}",new_x="LMARGIN",new_y="NEXT"); pdf.set_text_color(0,0,0)
        pdf.set_font("Helvetica","",9); pdf.s_cell(0,6,f"  {DESC.get(s['sc'],s['sc'])}",new_x="LMARGIN",new_y="NEXT"); pdf.ln(2)
        pdf.param_table(s["params"])
        car = CARACTERES.get(s['sc'])
        if car:
            pdf.sub_title("Caracteres principaux")
            pdf.set_font("Helvetica","I",8); pdf.s_multi(0,4,car); pdf.ln(2)
        if s.get("etats"):
            pdf.sub_title("Etat hydrique")
            pdf.set_font("Helvetica","B",9); pdf.s_cell(0,6,f"  Etat retenu : {s['etat']} ({ETQ.get(s['etat'],'')})",new_x="LMARGIN",new_y="NEXT")
            pdf.set_font("Helvetica","",8)
            for meth,et,val in s["etats"]: pdf.s_cell(0,4,f"    {meth} = {val} => {et} ({ETQ.get(et,'')})",new_x="LMARGIN",new_y="NEXT")
            if s.get("divergence"):
                pdf.set_font("Helvetica","I",7); pdf.set_text_color(170,51,51)
                pdf.s_cell(0,4,f"  {s['divergence']}",new_x="LMARGIN",new_y="NEXT")
                pdf.set_text_color(0,0,0)
            if s.get("insensible"): pdf.set_font("Helvetica","B",8); pdf.s_cell(0,5,"  Sol insensible a l'eau",new_x="LMARGIN",new_y="NEXT")

        # Remblai
        if s.get("do_remblai"):
            pdf.ln(3); pdf.chapter_title(f"{i}.2","Conditions d'utilisation en remblai")
            cond=s.get("remblai_cond")
            if isinstance(cond,str): pdf.set_font("Helvetica","B",9); pdf.s_multi(0,5,cond)
            else:
                for meteo,sols in cond.items():
                    pdf.set_font("Helvetica","B",8); pdf.s_cell(0,5,f"  {meteo}",new_x="LMARGIN",new_y="NEXT")
                    if isinstance(sols,str): pdf.set_font("Helvetica","",8); pdf.s_cell(0,4,f"    => {sols}",new_x="LMARGIN",new_y="NEXT")
                    else:
                        for sol_nom,codes in sols:
                            pdf.set_font("Helvetica","I",8); pdf.s_cell(0,4,f"    Solution : {sol_nom}",new_x="LMARGIN",new_y="NEXT")
                            pdf.set_x(18); pdf.code_row(codes)
                obs_txt = s.get("obs_remblai")
                if obs_txt:
                    pdf.set_font("Helvetica","B",8); pdf.s_cell(0,5,"  Observations (Fascicule 2) :",new_x="LMARGIN",new_y="NEXT")
                    pdf.set_font("Helvetica","I",8); pdf.s_multi(0,4,f"  {obs_txt}")

        # CdF
        if s.get("do_cdf"):
            pdf.ln(2); pdf.chapter_title(f"{i}.3","Couche de forme")
            cdf=s.get("cdf_result",{})
            pdf.set_font("Helvetica","",9); pdf.s_multi(0,5,cdf.get("texte",""))
            if cdf.get("traitement"): pdf.set_font("Helvetica","B",8); pdf.s_cell(0,5,f"  Traitement : {cdf['traitement']}",new_x="LMARGIN",new_y="NEXT")
            if cdf.get("params_meca"): pdf.param_table(cdf["params_meca"])
            if s.get("pst"):
                pdf.set_font("Helvetica","B",9); pdf.s_cell(0,6,f"  Classe PST simplifiee : {s['pst']}",new_x="LMARGIN",new_y="NEXT")

        # Compactage
        if s.get("do_compact"):
            pdf.ln(2); pdf.chapter_title(f"{i}.4","Compactage")
            pdf.set_font("Helvetica","B",8); pdf.set_fill_color(239,230,212)
            for h in ["Objectif","Ouvrage","rhodm","rhodfc"]: pdf.s_cell(35,5,h,border=1,fill=True)
            pdf.ln(); pdf.set_font("Helvetica","",8)
            for obj,ouv,dm,dfc in [("q4","Remblais/PST",">=95% rhodOPN",">=92% rhodOPN"),("q3","CdF",">=98,5% rhodOPN",">=96% rhodOPN")]:
                for v in [obj,ouv,dm,dfc]: pdf.s_cell(35,4,v,border=1)
                pdf.ln()
            pdf.ln(2)
            comp=s.get("compacteurs",{})
            if comp.get("q4"):
                pdf.set_font("Helvetica","B",8); pdf.s_cell(0,5,"  Remblai (q4):",new_x="LMARGIN",new_y="NEXT"); pdf.set_font("Helvetica","",8)
                for eng,d2 in comp["q4"]: pdf.s_cell(0,4,f"    - {eng}: {d2}",new_x="LMARGIN",new_y="NEXT")
            if comp.get("q3"):
                pdf.set_font("Helvetica","B",8); pdf.s_cell(0,5,"  CdF (q3):",new_x="LMARGIN",new_y="NEXT"); pdf.set_font("Helvetica","",8)
                for eng,d2 in comp["q3"]: pdf.s_cell(0,4,f"    - {eng}: {d2}",new_x="LMARGIN",new_y="NEXT")
            # Q/S table in PDF
            qs_data = get_qs_table(s.get("sc",""), s.get("famille",""))
            if qs_data:
                pdf.ln(3); pdf.sub_title("Tableau Q/S (Fascicule 2, Annexe 4)")
                for obj_label in ("q4", "q3"):
                    obj_data = qs_data.get(obj_label, {})
                    if obj_data:
                        pdf.set_font("Helvetica","B",7); pdf.s_cell(0,5,f"  {obj_label.upper()} :",new_x="LMARGIN",new_y="NEXT")
                        pdf.set_fill_color(239,230,212); pdf.set_font("Helvetica","B",7)
                        for h in ["Engin","Q/S (m3/m2)","e (m)","V (km/h)"]: pdf.s_cell(30,5,h,border=1,fill=True,align="C")
                        pdf.ln(); pdf.set_font("Helvetica","",7)
                        for eng, vals in obj_data.items():
                            pdf.s_cell(30,4,eng,border=1,align="C")
                            pdf.s_cell(30,4,f"{vals['qs']:.3f}",border=1,align="C")
                            pdf.s_cell(30,4,f"{vals['e']:.2f}",border=1,align="C")
                            pdf.s_cell(30,4,f"{vals['v']:.1f}",border=1,align="C")
                            pdf.ln()
                        pdf.ln(2)

    # SYNTHÈSE
    pdf.add_page(); pdf.chapter_title("","SYNTHESE GENERALE")
    pdf.set_font("Helvetica","B",8); pdf.set_fill_color(239,230,212)
    for h in ["Sondage","Profondeur","Classe","Etat","Remblai","CdF"]: pdf.s_cell(27,6,h,border=1,fill=True,align="C")
    pdf.ln(); pdf.set_font("Helvetica","",8)
    for s in sondages:
        depth_s = ""
        if s.get("prof_de") is not None and s.get("prof_a") is not None:
            depth_s = f"{s['prof_de']}-{s['prof_a']}m"
        pdf.s_cell(27,5,s["sondage_id"],border=1,align="C")
        pdf.s_cell(27,5,depth_s,border=1,align="C")
        pdf.s_cell(27,5,s["symbole"],border=1,align="C")
        pdf.s_cell(27,5,ETQ.get(s.get("etat",""),""),border=1,align="C")
        pdf.s_cell(27,5,"Oui" if s.get("do_remblai") and not isinstance(s.get("remblai_cond"),str) else "Etude",border=1,align="C")
        cdf=s.get("cdf_result",{})
        pdf.s_cell(27,5,"Oui" if "Utilisable" in cdf.get("texte","") else "Traitement" if cdf.get("traitement") else "-",border=1,align="C")
        pdf.ln()

    # COUPE GEOTECHNIQUE
    draw_coupe(pdf, sondages)

    pdf.ln(8); pdf.set_font("Helvetica","I",8); pdf.set_text_color(107,95,80)
    pdf.s_multi(0,4,"Ce rapport a ete genere par l'outil Classification GTR2023 developpe par Serigne Mouhamadane SY (UIDT). "
    "Il constitue un aide-memoire et ne se substitue pas a l'analyse d'un geotechnicien qualifie ni a la lecture du GTR 2023 (IDRRIM/Cerema).")

    buf=BytesIO(); pdf.output(buf); buf.seek(0); return buf

# ══════════════════════════════════════════════════════════════════
# SESSION
# ══════════════════════════════════════════════════════════════════
if "step" not in st.session_state: st.session_state.step=0
if "sondages" not in st.session_state: st.session_state.sondages=[]
if "projet_info" not in st.session_state: st.session_state.projet_info={}

def go(n): st.session_state.step=n
def reset_all():
    for k in list(st.session_state.keys()): del st.session_state[k]
    st.session_state.step=0; st.session_state.sondages=[]; st.session_state.projet_info={}
def new_sondage():
    st.session_state.step=1
    if "cur" in st.session_state: del st.session_state["cur"]

def save_project():
    data={"projet_info":st.session_state.projet_info,"sondages":st.session_state.sondages,
          "date_sauvegarde":datetime.now().strftime("%d/%m/%Y %H:%M")}
    return json.dumps(data, ensure_ascii=False, indent=2)

def load_project(content):
    data=json.loads(content)
    st.session_state.projet_info=data["projet_info"]
    st.session_state.sondages=data.get("sondages",[])
    if st.session_state.sondages:
        st.session_state.step=7
    else:
        st.session_state.step=1

show_header()

# ══════════════════════════════════════════════════════════════════
# ÉTAPE 0 — ACCUEIL PROJET
# ══════════════════════════════════════════════════════════════════
if st.session_state.step==0:
    st.title("Classification GTR2023")
    st.markdown("### Bienvenue")

    tab_new, tab_open = st.tabs(["📄 Nouveau projet", "📂 Ouvrir un projet"])

    with tab_new:
        st.markdown("Renseignez les informations du projet pour commencer.")
        projet=st.text_input("Nom du projet",placeholder="Ex : Autoroute Ila Touba — Lot 3")
        site=st.text_input("Nom du site",placeholder="Ex : PK 12+500 à PK 15+000")
        nb_sondages=st.number_input("Nombre de sondages prévus",1,100,value=1)
        ingenieur=st.text_input("Nom de l'ingénieur",placeholder="Ex : M. Diallo, Ing. géotechnicien")
        st.divider()
        if st.button("▶ Commencer un nouveau projet",type="primary",use_container_width=True):
            st.session_state.projet_info={"projet":projet or "","site":site or "","nb_sondages":int(nb_sondages),"ingenieur":ingenieur or ""}
            go(1); st.rerun()

    with tab_open:
        st.markdown("Chargez un projet sauvegardé précédemment (fichier `.json`).")
        uploaded = st.file_uploader("Sélectionner le fichier projet", type=["json"], label_visibility="collapsed")
        if uploaded is not None:
            try:
                content = uploaded.read().decode("utf-8")
                data = json.loads(content)
                pi = data.get("projet_info", {})
                sondages = data.get("sondages", [])
                date_sauv = data.get("date_sauvegarde", "")

                st.success(f"✅ Projet chargé : **{pi.get('projet', 'Sans nom')}**")
                st.markdown(f"- **Site** : {pi.get('site', '-')}")
                st.markdown(f"- **Ingénieur** : {pi.get('ingenieur', '-')}")
                st.markdown(f"- **Sondages enregistrés** : {len(sondages)}")
                if date_sauv:
                    st.markdown(f"- **Sauvegardé le** : {date_sauv}")

                if sondages:
                    with st.expander(f"📋 {len(sondages)} sondage(s) dans le projet"):
                        for s in sondages:
                            st.markdown(f"• **{s['sondage_id']}** → {s['symbole']} ({DESC.get(s['sc'], s['sc'])})")

                if st.button("▶ Ouvrir ce projet", type="primary", use_container_width=True):
                    load_project(content)
                    st.rerun()
            except Exception as e:
                st.error(f"❌ Fichier invalide : {e}")

    st.markdown(CREDIT, unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════
# ÉTAPE 1 — SAISIE SONDAGE + CLASSIFICATION
# ══════════════════════════════════════════════════════════════════
elif st.session_state.step==1:
    pi=st.session_state.projet_info
    n_done=len(st.session_state.sondages)
    st.markdown(f"### Sondage {n_done+1} / {pi['nb_sondages']}")
    st.caption(f"Projet : {pi['projet']} — Site : {pi['site']}")
    if n_done>0:
        with st.expander(f"📋 {n_done} sondage(s) déjà enregistré(s)"):
            for s in st.session_state.sondages:
                st.markdown(f"• **{s['sondage_id']}** → {s['symbole']} ({DESC.get(s['sc'],s['sc'])})")
    st.divider()
    _prefill_sid = st.session_state.pop("_prefill_sid", "")
    _prefill_prof_de = st.session_state.pop("_prefill_prof_de", 0.0)
    sondage_id=st.text_input("Numéro / Identifiant du sondage *",value=_prefill_sid,placeholder="Ex : S1, BH-03, SP-12...")

    st.markdown("**Profondeur du sondage**")
    c_prof1, c_prof2 = st.columns(2)
    with c_prof1: prof_de = st.number_input("De (m)", 0.0, 100.0, value=float(_prefill_prof_de), step=0.5, key="prof_de")
    with c_prof2: prof_a = st.number_input("A (m)", 0.0, 100.0, value=None, placeholder="Requis", step=0.5, key="prof_a")

    mat_type=st.radio("Type de matériau *",["Sol meuble (Dmax ≤ 63 mm)","Sol à gros éléments (Dmax > 63 mm) — VC","Matériau rocheux extrait","Sol organique (MO > 2 %)"])
    st.divider()
    errors=[]

    if "Sol meuble" in mat_type:
        st.markdown("**Granulométrie** *(NF EN ISO 17892-4)*")
        c1,c2,c3=st.columns(3)
        with c1: p63=st.number_input("Passant 63µm (%) *",0.0,100.0,value=None,placeholder="Obligatoire")
        with c2: p2=st.number_input("Passant 2mm (%)",0.0,100.0,value=None,placeholder="Si p63<15%",help="Obligatoire si passant 63µm < 15% (distinction S/G). Optionnel sinon (optionnel).")
        with c3: cu=st.number_input("Cu (D60/D10)",0.0,500.0,value=None,placeholder="Si disponible",help="Utilisé pour S1/S2 vs S3/S4 (seuil 6). Optionnel (optionnel).")
        st.markdown("**Argilosité** *(au moins IP ou VBS — les deux si disponibles)*")
        c4,c5=st.columns(2)
        with c4: ip=st.number_input("IP (%) *",0.0,100.0,value=None,placeholder="IP ou VBS requis",help="NF EN ISO 17892-12 — Seuils : 12, 22, 40, 55. Prioritaire pour sols moyennement à très argileux.")
        with c5: vbs=st.number_input("VBS (g/100g) *",0.0,20.0,value=None,placeholder="IP ou VBS requis",help="NF EN 17542-3 — Seuils : 1.5, 2.5, 6, 8. Prioritaire pour sols F1, I et S/G.")
        st.markdown("**État hydrique** *(optionnel — au moins wn+wOPN ou IPI recommandé)*")
        c6,c7,c8=st.columns(3)
        with c6: wn=st.number_input("wn (%)",0.0,200.0,value=None,placeholder="Optionnel",help="Teneur en eau naturelle. Utilisé avec wOPN pour le ratio wn/wOPN (optionnel).")
        with c7: wopn=st.number_input("wOPN (%)",0.0,100.0,value=None,placeholder="Optionnel",help="Optimum Proctor Normal. Fiable pour les états s et ts (optionnel).")
        with c8: ipi=st.number_input("IPI",0.0,100.0,value=None,placeholder="Optionnel",help="NF EN 13286-47 — Privilégié pour les états h et th (optionnel).")
        use_ic=st.checkbox("Utiliser Ic (recommandé pour F2, F3, F4, I2)")
        ic=st.number_input("Ic",0.0,3.0,value=None,placeholder="Optionnel") if use_ic else None

        if st.button("✅ Classifier",type="primary",use_container_width=True):
            if not sondage_id: errors.append("N° sondage")
            if prof_a is None: errors.append("Profondeur A (m)")
            if p63 is None: errors.append("Passant 63 µm (obligatoire)")
            if ip is None and vbs is None: errors.append("IP ou VBS (au moins un des deux)")
            if p63 is not None and p63<15 and p2 is None: errors.append("Passant 2 mm (obligatoire si passant 63µm < 15%)")
            if errors: st.error(f"⚠ {' — '.join(errors)}")
            else:
                fam,sc=classify(p63,p2,ip,vbs,cu); etats=calc_etat(sc,wn,wopn,ipi,ic,p2); etat,divergence_msg=resolve_etat(etats)
                if fam in ("S","G") and cu is None:
                    st.warning(f"⚠ Cu non renseigné — le sol est classé avec Cu < 6 par défaut ({sc}). Pour une classification précise des sols S et G, le Cu est recommandé.")
                params={"Prof. de":f"{prof_de}m","Prof. a":f"{prof_a}m","Pass. 63µm":f"{p63:.1f}%"}
                if p2 is not None: params["Pass. 2mm"]=f"{p2:.1f}%"
                if cu is not None: params["Cu"]=f"{cu:.1f}"
                if ip is not None: params["IP"]=f"{ip:.1f}%"
                if vbs is not None: params["VBS"]=f"{vbs:.2f}"
                if wn is not None: params["wn"]=f"{wn:.1f}%"
                if wopn is not None: params["wOPN"]=f"{wopn:.1f}%"
                if ipi is not None: params["IPI"]=f"{ipi:.1f}"
                if ic is not None: params["Ic"]=f"{ic:.2f}"
                st.session_state.cur={"type":"sol","sondage_id":sondage_id,"famille":fam,"sc":sc,"etat":etat,"etats":etats,"insensible":False,
                    "symbole":sc+etat,"params":params,"p63":p63,"p2":p2,"vbs_val":vbs,"divergence":divergence_msg,
                    "prof_de":prof_de,"prof_a":prof_a}
                go(2); st.rerun()

    elif "gros éléments" in mat_type:
        c1,c2=st.columns(2)
        with c1: dmax=st.number_input("Dmax (mm)*",64.0,5000.0,value=None,placeholder="Requis")
        with c2: frac=st.number_input("Frac. 0/63mm (%)*",0.0,100.0,value=None,placeholder="Requis")
        vc_t=st.radio("Type",["VC1 — Très charpentés","VC2 — Peu charpentés"]); vc="VC1" if "VC1" in vc_t else "VC2"
        st.markdown("**Fraction 0/63mm :**")
        c3,c4,c5=st.columns(3)
        with c3: p63v=st.number_input("Pass. 63µm (%) *",0.0,100.0,value=None,placeholder="Obligatoire",key="vp63")
        with c4: p2v=st.number_input("Pass. 2mm (%)",0.0,100.0,value=None,placeholder="Si p63<15%",key="vp2",help="Obligatoire si passant 63µm < 15% (optionnel).")
        with c5: cuv=st.number_input("Cu",0.0,500.0,value=None,placeholder="Si disponible",key="vcu",help="Optionnel (optionnel).")
        st.markdown("**Argilosité** *(au moins IP ou VBS)*")
        c6,c7=st.columns(2)
        with c6: ipv=st.number_input("IP (%) *",0.0,100.0,value=None,placeholder="IP ou VBS requis",key="vip")
        with c7: vbsv=st.number_input("VBS *",0.0,20.0,value=None,placeholder="IP ou VBS requis",key="vvbs")
        if st.button("✅ Classifier",type="primary",use_container_width=True):
            if not sondage_id: errors.append("N° sondage")
            if prof_a is None: errors.append("Profondeur A (m)")
            if dmax is None: errors.append("Dmax")
            if frac is None: errors.append("Fraction 0/63")
            if p63v is None: errors.append("Passant 63µm")
            if p63v is not None and p63v<15 and p2v is None: errors.append("Passant 2mm (obligatoire si p63<15%)")
            if ipv is None and vbsv is None: errors.append("IP ou VBS (au moins un)")
            if errors: st.error(f"⚠ {' — '.join(errors)}")
            else:
                f0,sc0=classify(p63v,p2v,ipv,vbsv,cuv)
                vc_params={"Prof. de":f"{prof_de}m","Prof. a":f"{prof_a}m","Dmax":f"{dmax:.0f}mm","Frac.0/63":f"{frac:.0f}%","Type":vc,"P63µm(0/63)":f"{p63v:.1f}%","P2mm(0/63)":f"{p2v:.1f}%"}
                if cuv is not None: vc_params["Cu(0/63)"]=f"{cuv:.1f}"
                if ipv is not None: vc_params["IP(0/63)"]=f"{ipv:.1f}%"
                if vbsv is not None: vc_params["VBS(0/63)"]=f"{vbsv:.2f}"
                st.session_state.cur={"type":"vc","sondage_id":sondage_id,"famille":"VC","sc":f"{vc}({sc0})","scm":f"{vc}({sc0})",
                    "etat":None,"symbole":f"{vc}({sc0})","etats":[],"insensible":False,
                    "params":vc_params,"p2":p2v,"vc_sc063":sc0,"vc_fam063":f0,"do_remblai":False,"do_cdf":False,"do_compact":False,
                    "prof_de":prof_de,"prof_a":prof_a}
                go(2); st.rerun()

    elif "rocheux" in mat_type:
        roche=st.selectbox("Classe*",list(ROCHES.keys()),format_func=lambda k:f"{k} — {ROCHES[k]}")
        c1,c2=st.columns(2)
        with c1: la_r=st.number_input("LA*",0,100,value=None,placeholder="Requis")
        with c2: mde_r=st.number_input("MDE*",0,100,value=None,placeholder="Requis")
        c3,c4=st.columns(2)
        with c3: ifr=st.number_input("IFR",0.0,50.0,value=None,placeholder="Optionnel")
        with c4: idga=st.number_input("IDGa",0.0,50.0,value=None,placeholder="Optionnel")
        rho_d=wn_ch=None
        if roche=="CH":
            c5,c6=st.columns(2)
            with c5: rho_d=st.number_input("ρd (t/m³)*",0.5,3.0,value=None,placeholder="Requis")
            with c6: wn_ch=st.number_input("wn (%)*",0.0,60.0,value=None,placeholder="Requis")
        if st.button("✅ Classifier",type="primary",use_container_width=True):
            if not sondage_id: errors.append("N° sondage")
            if prof_a is None: errors.append("Profondeur A (m)")
            if la_r is None: errors.append("LA");
            if mde_r is None: errors.append("MDE")
            if roche=="CH":
                if rho_d is None: errors.append("ρd")
                if wn_ch is None: errors.append("wn")
            if errors: st.error(f"Champs obligatoires : {', '.join(errors)}")
            else:
                params={"Prof. de":f"{prof_de}m","Prof. a":f"{prof_a}m","Classe":roche,"Nature":ROCHES[roche],"LA":str(la_r),"MDE":str(mde_r)}
                if ifr is not None: params["IFR"]=f"{ifr:.1f}"
                if idga is not None: params["IDGa"]=f"{idga:.1f}"
                sc_r=roche
                if roche=="CH":
                    params["ρd"]=f"{rho_d:.2f}"; params["wn"]=f"{wn_ch:.1f}%"
                    if rho_d>1.95: sc_r="CH1"
                    elif rho_d>1.70: sc_r="CH2"
                    elif rho_d>1.55: sc_r="CH3"+("h" if wn_ch>=27 else "m" if wn_ch>=22 else "s" if wn_ch>=18 else "ts")
                    else: sc_r="CH4"+("th" if wn_ch>=31 else "h" if wn_ch>=26 else "m" if wn_ch>=21 else "s" if wn_ch>=16 else "ts")
                elif roche in ("Cl","Sa","Co","Vo","Me"):
                    # Sub-classification R3/R4/R5 based on IFR and IDGa
                    if ifr is not None and idga is not None:
                        if ifr > 14 or idga > 20: sc_r = roche + "-R5"
                        elif ifr > 7 or idga > 5: sc_r = roche + "-R4"
                        else: sc_r = roche + "-R3"
                    elif ifr is not None:
                        if ifr > 14: sc_r = roche + "-R5"
                        elif ifr > 7: sc_r = roche + "-R4"
                        else: sc_r = roche + "-R3"
                    elif idga is not None:
                        if idga > 20: sc_r = roche + "-R5"
                        elif idga > 5: sc_r = roche + "-R4"
                        else: sc_r = roche + "-R3"
                cdf_ok=la_r<=45 and mde_r<=45
                st.session_state.cur={"type":"roche","sondage_id":sondage_id,"famille":"R","sc":sc_r,"scm":sc_r,"etat":None,"symbole":sc_r,
                    "insensible":False,"etats":[],"params":params,"p2":None,"la":la_r,"mde":mde_r,"prof_de":prof_de,"prof_a":prof_a,
                    "do_remblai":True,"remblai_cond":"Matériau rocheux — Fascicule 2, Annexe 2.",
                    "do_cdf":True,"cdf_result":{"texte":f"LA={la_r}, MDE={mde_r} — {'Utilisable' if cdf_ok else 'Non utilisable'}",
                    "params_meca":{"LA":str(la_r),"MDE":str(mde_r)},"traitement":"" if cdf_ok else "Traitement/substitution"},
                    "do_compact":True,"compacteurs":{"q4":[("V4-V5","Vibrant lisse"),("P3","Pneus")],"q3":[("V5","Vibrant lisse")]}}
                go(2); st.rerun()
    else:
        mo=st.number_input("MO (%)*",0.0,100.0,value=None,placeholder="Requis")
        if st.button("✅ Classifier",type="primary",use_container_width=True):
            if not sondage_id: errors.append("N° sondage")
            if prof_a is None: errors.append("Profondeur A (m)")
            if mo is None: errors.append("MO")
            if errors: st.error(f"Champs obligatoires : {', '.join(errors)}")
            else:
                sc_o="Non organique" if mo<2 else "O1" if mo<6 else "O2" if mo<20 else "O3"
                st.session_state.cur={"type":"organique","sondage_id":sondage_id,"famille":"O","sc":sc_o,"scm":sc_o,"etat":None,
                    "symbole":sc_o,"insensible":False,"etats":[],"params":{"Prof. de":f"{prof_de}m","Prof. a":f"{prof_a}m","MO":f"{mo:.1f}%"},"p2":None,
                    "do_remblai":False,"do_cdf":False,"do_compact":False,"prof_de":prof_de,"prof_a":prof_a}
                go(2); st.rerun()

# ══════════════════════════════════════════════════════════════════
# ÉTAPE 2 — RÉSULTAT CLASSIFICATION
# ══════════════════════════════════════════════════════════════════
elif st.session_state.step==2:
    d=st.session_state.cur
    st.markdown(f"### Résultat — Sondage {d['sondage_id']}")
    desc=DESC.get(d["sc"],ROCHES.get(d["sc"],d["sc"])) if d["type"] not in ("organique",) else {"O1":"Faible MO (2-6%)","O2":"MO modérée (6-20%)","O3":"Tourbe (>20%)","Non organique":"MO<2%"}.get(d["sc"],d["sc"])
    el=ETQ.get(d["etat"],"") if d["etat"] else ""
    st.markdown(f'<div class="result-box"><div class="class-symbol">{d["symbole"]}</div><div style="font-size:14px;color:#6b5f50;margin-top:4px;">{desc}</div>'
        f'<div style="margin-top:10px;"><span class="badge badge-main">{d["famille"]}</span><span class="badge">{d["sc"]}</span>'
        f'{"<span class=badge>"+el+"</span>" if el else ""}</div></div>',unsafe_allow_html=True)
    cols=st.columns(3)
    for i,(k,v) in enumerate(d["params"].items()): cols[i%3].markdown(f"• **{k}**: {v}")
    if d["etats"]:
        st.markdown("**État hydrique :**")
        for meth,et,val in d["etats"]: st.markdown(f"• {meth}={val} → **{et}** ({ETQ.get(et,'')})")
        if d.get("divergence"):
            st.warning(f"⚠️ {d['divergence']}")
    car = CARACTERES.get(d["sc"])
    if car:
        st.markdown("**Caractères principaux :**")
        st.markdown(f"*{car}*")
    # Depth display
    if d.get("prof_de") is not None and d.get("prof_a") is not None:
        st.markdown(f"**Profondeur :** {d['prof_de']} - {d['prof_a']} m")
    # Synoptique for sol meuble
    if d["type"]=="sol":
        show_synoptique(d["famille"], d["sc"], d.get("p63"))
    st.divider()

    if d["type"]=="organique":
        if d["sc"]=="O1": st.markdown("• **Remblai**: ✅ Oui (H<10m)\n• **CdF**: ✅ Avec étude si MO > 3 %")
        elif d["sc"]=="O2": st.markdown("• **Remblai**: ⚠️ Merlons, surfaces enherbées\n• **CdF**: ❌ Non")
        elif d["sc"]=="O3": st.markdown("• **Remblai**: ❌ Non\n• **CdF**: ❌ Non")
        else: st.success("MO < 2 % — Classifier comme sol meuble ou roche.")
        d["do_remblai"]=False;d["do_cdf"]=False;d["do_compact"]=False
        if st.button("💾 Enregistrer ce sondage",type="primary",use_container_width=True):
            st.session_state.sondages.append(d); go(7); st.rerun()
    elif d["type"]=="vc":
        st.info(f"Sol VC — le comportement est contrôlé par la fraction 0/63 mm (classe {d.get('vc_sc063','')}).")
        st.markdown("**Souhaitez-vous détailler les conditions d'utilisation ?**")
        c1,c2=st.columns(2)
        with c1:
            if st.button("▶ Continuer (remblai, CdF, compactage)",type="primary",use_container_width=True):
                d["famille"]=d.get("vc_fam063","G"); d["sc_orig"]=d["sc"]; d["sc"]=d.get("vc_sc063","G1")
                go(3 if d["sc"][0] in "SG" else 4); st.rerun()
        with c2:
            if st.button("💾 Enregistrer sans détail",use_container_width=True):
                d["do_remblai"]=False;d["do_cdf"]=False;d["do_compact"]=False
                st.session_state.sondages.append(d); go(7); st.rerun()
    elif d["type"]=="roche":
        la_v=d.get("la",99); mde_v=d.get("mde",99)
        st.markdown("**Aperçu :**")
        if d["sc"].startswith("SR"): st.error("⚠️ Roches salines — études spécifiques obligatoires.")
        elif la_v<=45 and mde_v<=45: st.success(f"✅ LA={la_v} ≤ 45, MDE={mde_v} ≤ 45 — Utilisable en remblai et CdF.")
        else: st.warning(f"⚠️ LA={la_v} ou MDE={mde_v} > 45 — CdF non utilisable sans traitement.")
        st.markdown("**Souhaitez-vous détailler les conditions d'utilisation ?**")
        c1,c2=st.columns(2)
        with c1:
            if st.button("▶ Continuer (remblai, CdF, compactage)",type="primary",use_container_width=True):
                cdf_ok=la_v<=45 and mde_v<=45
                d["do_remblai"]=True;d["remblai_cond"]="Matériau rocheux — se reporter au Fascicule 2, Annexe 2."
                d["do_cdf"]=True;d["cdf_result"]={"texte":f"LA={la_v}, MDE={mde_v} — {'Utilisable' if cdf_ok else 'Non utilisable'}",
                    "params_meca":{"LA":str(la_v),"MDE":str(mde_v)},"traitement":"" if cdf_ok else "Traitement/substitution"}
                d["do_compact"]=True;d["compacteurs"]={"q4":[("V4-V5","Vibrant lisse"),("P3","Pneus")],"q3":[("V5","Vibrant lisse")]}
                st.session_state.sondages.append(d); go(7); st.rerun()
        with c2:
            if st.button("💾 Enregistrer sans détail",use_container_width=True):
                d["do_remblai"]=False;d["do_cdf"]=False;d["do_compact"]=False
                st.session_state.sondages.append(d); go(7); st.rerun()
    else:
        st.markdown("**Souhaitez-vous continuer vers les conditions d'utilisation ?**")
        c1,c2=st.columns(2)
        with c1:
            if st.button("▶ Continuer (remblai, CdF, compactage)",type="primary",use_container_width=True):
                go(3 if d["sc"][0] in "SG" else 4); st.rerun()
        with c2:
            if st.button("💾 Enregistrer sans détail",use_container_width=True):
                d["do_remblai"]=False;d["do_cdf"]=False;d["do_compact"]=False
                st.session_state.sondages.append(d); go(7); st.rerun()

# ÉTAPE 3 — SENSIBILITÉ EAU (S/G)
elif st.session_state.step==3:
    d=st.session_state.cur
    st.markdown(f"### Sensibilité à l'eau — {d['sondage_id']} ({d['sc']})")
    cbri=st.number_input("CBRi (après 4j immersion)*",0.0,100.0,value=None,placeholder="Requis")
    if st.button("✅ Vérifier",type="primary",use_container_width=True):
        if cbri is None: st.error("CBRi obligatoire")
        else:
            ins=check_ins(d["sc"],d["p63"],d["vbs_val"],cbri); d["insensible"]=ins; d["params"]["CBRi"]=f"{cbri:.0f}"
            if ins: d["etat"]="ins"; d["symbole"]=d["sc"]+"ins"
            go(4); st.rerun()

# ÉTAPE 4 — REMBLAI
elif st.session_state.step==4:
    d=st.session_state.cur
    st.markdown(f"### Remblai — {d['sondage_id']} ({d['symbole']})")
    if d.get("insensible"): st.success("✅ Insensible à l'eau.")
    cond=get_remblai(d["sc"],d["etat"],d["famille"]); d["remblai_cond"]=cond; d["do_remblai"]=True
    if isinstance(cond,str): st.error(f"❌ {cond}")
    else:
        for meteo,sols in cond.items():
            with st.expander(f"**{meteo}**",expanded="(=)" in meteo):
                if isinstance(sols,str): st.error(sols)
                else:
                    for sol_nom,codes in sols:
                        code_str=" ".join(str(codes.get(x,0)) for x in ["E","G","W","T","R","C","H"])
                        st.markdown(f"*{sol_nom}*"); st.markdown(f'<div class="code-box">E G W T R C H → {code_str}</div>',unsafe_allow_html=True)
                        for rub,val in codes.items():
                            if val!=0: st.markdown(f"• **{rub}{val}**: {CODES[rub][val]}")
    # Observations remblai par sous-classe
    obs_sc = OBS_REMBLAI.get(d["sc"], {})
    obs_txt = obs_sc.get(d.get("etat", "m"))
    if obs_txt:
        st.markdown("**Observations (Fascicule 2) :**")
        st.info(obs_txt)
        d["obs_remblai"] = obs_txt
    st.divider()
    c1,c2=st.columns(2)
    with c1:
        if st.button("▶ Couche de forme",type="primary",use_container_width=True): go(5); st.rerun()
    with c2:
        if st.button("⏩ Passer au compactage",use_container_width=True): d["do_cdf"]=False; go(6); st.rerun()

# ÉTAPE 5 — COUCHE DE FORME
elif st.session_state.step==5:
    d=st.session_state.cur; fam=d["famille"]
    st.markdown(f"### Couche de forme — {d['sondage_id']} ({d['symbole']})")
    errors=[]
    if fam=="S":
        fs=st.number_input("FS — Friabilité des sables*",0,100,value=None,placeholder="Requis")
        if st.button("✅ Valider",type="primary",use_container_width=True):
            if fs is None: st.error("FS obligatoire")
            else:
                scm=d["sc"]+("1" if fs<=60 else "2"); d["scm"]=scm; d["do_cdf"]=True
                texte=f"Utilisable ({scm}, FS={fs}≤60)" if fs<=60 else f"Insuffisant ({scm}, FS={fs}>60)"
                d["cdf_result"]={"texte":texte,"params_meca":{"FS":str(fs)},"traitement":"" if fs<=60 else "Traitement LH"}; d["params"]["FS"]=str(fs)
                go(6); st.rerun()
    elif fam=="G":
        c1,c2=st.columns(2)
        with c1: la=st.number_input("LA*",0,100,value=None,placeholder="Requis")
        with c2: mde=st.number_input("MDE*",0,100,value=None,placeholder="Requis")
        if st.button("✅ Valider",type="primary",use_container_width=True):
            if la is None: errors.append("LA")
            if mde is None: errors.append("MDE")
            if errors: st.error(f"Obligatoire: {', '.join(errors)}")
            else:
                scm=d["sc"]+("1" if la<=45 and mde<=45 else "2"); d["scm"]=scm; d["do_cdf"]=True
                texte=f"Utilisable ({scm})" if la<=45 and mde<=45 else f"Insuffisant ({scm})"
                d["cdf_result"]={"texte":texte,"params_meca":{"LA":str(la),"MDE":str(mde)},"traitement":"" if la<=45 and mde<=45 else "Traitement LH/substitution"}
                d["params"].update({"LA":str(la),"MDE":str(mde)}); go(6); st.rerun()
    else:
        d["scm"]=d["sc"]; d["do_cdf"]=True
        texte="CdF avec traitement chaux/LH" if d["sc"] not in ("F4","F4+") else "Non recommandé — étude spécifique"
        d["cdf_result"]={"texte":texte,"params_meca":{},"traitement":"Chaux et/ou LH (cf. GTS)"}
        st.info(texte)
        if st.button("▶ Compactage",type="primary",use_container_width=True): go(6); st.rerun()
    # PST determination
    pst_val, pst_note = get_pst(d.get("sc",""), d.get("etat","m"), d.get("famille",""))
    if pst_val:
        st.markdown(f"**Classe PST simplifiée : {pst_val}**")
        d["pst"] = pst_val
        if pst_note: st.caption(pst_note)
    elif pst_note:
        st.caption(f"PST : {pst_note}")

# ÉTAPE 6 — COMPACTAGE
elif st.session_state.step==6:
    d=st.session_state.cur
    st.markdown(f"### Compactage — {d['sondage_id']} ({d['symbole']})")
    d["compacteurs"]=get_compacteurs(d["famille"],d["etat"]); d["do_compact"]=True
    st.markdown("|Objectif|Ouvrage|ρdm|ρdfc|\n|---|---|---|---|\n|**q4**|Remblais/PST|≥95% ρdOPN|≥92% ρdOPN|\n|**q3**|CdF|≥98,5% ρdOPN|≥96% ρdOPN|")
    st.markdown("**Remblai (q4):**")
    for eng,desc in d["compacteurs"]["q4"]: st.markdown(f"• **{eng}** — {desc}")
    st.markdown("**CdF (q3):**")
    for eng,desc in d["compacteurs"]["q3"]: st.markdown(f"• **{eng}** — {desc}")
    # Q/S table
    qs_data = get_qs_table(d["sc"], d["famille"])
    if qs_data:
        st.markdown("---")
        st.markdown("**Tableau Q/S de compactage** *(GTR 2023, Fascicule 2, Annexe 4)*")
        for obj_label in ("q4", "q3"):
            obj_data = qs_data.get(obj_label, {})
            if obj_data:
                st.markdown(f"***{obj_label.upper()} :***")
                rows = []
                for eng, vals in obj_data.items():
                    rows.append(f"| {eng} | {vals['qs']:.3f} | {vals['e']:.2f} | {vals['v']:.1f} |")
                table = "| Engin | Q/S (m³/m²) | e (m) | V (km/h) |\n|---|---|---|---|\n" + "\n".join(rows)
                st.markdown(table)
    st.divider()
    if st.button("💾 Enregistrer ce sondage",type="primary",use_container_width=True):
        st.session_state.sondages.append(d); go(7); st.rerun()

# ══════════════════════════════════════════════════════════════════
# ÉTAPE 7 — AJOUTER SONDAGE OU RAPPORT
# ══════════════════════════════════════════════════════════════════
elif st.session_state.step==7:
    pi=st.session_state.projet_info; sondages=st.session_state.sondages
    st.markdown(f"### ✅ {len(sondages)} sondage(s) enregistré(s)")
    st.caption(f"Projet : {pi['projet']} — Site : {pi['site']}")
    # Group sondages by sondage_id for display
    from collections import OrderedDict
    sondage_groups = OrderedDict()
    for idx, s in enumerate(sondages):
        sid = s["sondage_id"]
        if sid not in sondage_groups:
            sondage_groups[sid] = []
        sondage_groups[sid].append((idx, s))
    for sid, layers in sondage_groups.items():
        st.markdown(f"**{sid}**" + (f" ({len(layers)} couche(s))" if len(layers) > 1 else ""))
        for idx, s in layers:
            desc=DESC.get(s["sc"],s["sc"])
            depth_str = ""
            if s.get("prof_de") is not None and s.get("prof_a") is not None:
                depth_str = f" ({s['prof_de']}-{s['prof_a']} m)"
            c_s, c_del = st.columns([8, 1])
            with c_s:
                st.markdown(f'<div class="result-box" style="padding:12px;"><strong>{s["sondage_id"]}{depth_str}</strong> → '
                    f'<span class="badge badge-main">{s["symbole"]}</span> <span style="color:#6b5f50;font-size:13px;">{desc}</span></div>',unsafe_allow_html=True)
            with c_del:
                if st.button("🗑", key=f"del_{idx}", help=f"Supprimer {s['sondage_id']}{depth_str}"):
                    st.session_state.sondages.pop(idx); st.rerun()
    st.divider()
    remaining=pi["nb_sondages"]-len(sondages)
    if remaining>0:
        st.info(f"Il reste **{remaining} sondage(s)** à saisir sur les {pi['nb_sondages']} prévus.")
    c1,c2,c3=st.columns(3)
    with c1:
        if st.button("➕ Ajouter un sondage",type="primary",use_container_width=True):
            new_sondage(); st.rerun()
    with c2:
        if st.button("📥 Générer le rapport PDF",use_container_width=True):
            go(8); st.rerun()
    with c3:
        proj_json = save_project()
        nom_projet = pi.get("projet","projet")[:30].replace(" ","_") or "projet"
        st.download_button("💾 Sauvegarder le projet", data=proj_json,
            file_name=f"GTR2023_{nom_projet}_{datetime.now().strftime('%Y%m%d')}.json",
            mime="application/json", use_container_width=True)
    if sondages:
        last_s = sondages[-1]
        last_prof_a = last_s.get("prof_a", 0) or 0
        if st.button("➕ Ajouter une couche au sondage actuel", use_container_width=True):
            st.session_state.step = 1
            if "cur" in st.session_state: del st.session_state["cur"]
            st.session_state["_prefill_sid"] = last_s["sondage_id"]
            st.session_state["_prefill_prof_de"] = last_prof_a
            st.rerun()
    if sondages:
        xl_buf = build_excel(pi, sondages)
        nom_xl = pi.get("projet","projet")[:20].replace(" ","_") or "projet"
        st.download_button("📊 Exporter en Excel", data=xl_buf,
            file_name=f"GTR2023_{nom_xl}_{datetime.now().strftime('%Y%m%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)

# ══════════════════════════════════════════════════════════════════
# ÉTAPE 8 — RAPPORT FINAL
# ══════════════════════════════════════════════════════════════════
elif st.session_state.step==8:
    pi=st.session_state.projet_info; sondages=st.session_state.sondages
    show_header()
    st.markdown("<h2 style='text-align:center;color:#2e2620!important;'>RAPPORT GTR 2023</h2>",unsafe_allow_html=True)
    st.markdown(f"<p style='text-align:center;color:#6b5f50;font-size:12px;'>{pi['projet']} — {pi['site']} — {pi['ingenieur']} — {datetime.now().strftime('%d/%m/%Y')}</p>",unsafe_allow_html=True)

    st.markdown("### Synthèse")
    cols_h=st.columns([2,1,2,1,1,1])
    cols_h[0].markdown("**Sondage**"); cols_h[1].markdown("**Prof.**"); cols_h[2].markdown("**Classe**"); cols_h[3].markdown("**État**"); cols_h[4].markdown("**Remblai**"); cols_h[5].markdown("**CdF**")
    for s in sondages:
        c=st.columns([2,1,2,1,1,1])
        depth_disp = f"{s.get('prof_de',0)}-{s.get('prof_a','')}m" if s.get("prof_a") is not None else "-"
        c[0].markdown(s["sondage_id"]); c[1].markdown(depth_disp); c[2].markdown(f"**{s['symbole']}**"); c[3].markdown(ETQ.get(s.get("etat",""),"-"))
        c[4].markdown("✅" if s.get("do_remblai") and not isinstance(s.get("remblai_cond"),str) else "❌")
        cdf=s.get("cdf_result",{}); c[5].markdown("✅" if "Utilisable" in cdf.get("texte","") else "⚠️" if cdf.get("traitement") else "-")

    for s in sondages:
        st.divider()
        depth_str = ""
        if s.get("prof_de") is not None and s.get("prof_a") is not None:
            depth_str = f" ({s['prof_de']}-{s['prof_a']} m)"
        st.markdown(f"### Sondage {s['sondage_id']}{depth_str} — {s['symbole']}")
        st.markdown(f"*{DESC.get(s['sc'],s['sc'])}*")
        cols=st.columns(3)
        for i,(k,v) in enumerate(s["params"].items()): cols[i%3].markdown(f"• **{k}**: {v}")
        if s.get("do_remblai"):
            st.markdown("**Remblai :**")
            cond=s["remblai_cond"]
            if isinstance(cond,str): st.info(cond)
            else:
                for meteo,sols in cond.items():
                    if isinstance(sols,str): st.markdown(f"• {meteo}: ❌ {sols}")
                    else:
                        for sn,codes in sols:
                            cs=" ".join(str(codes.get(x,0)) for x in ["E","G","W","T","R","C","H"])
                            st.markdown(f"• {meteo} — *{sn}* : `{cs}`")
        if s.get("do_cdf"): st.markdown(f"**CdF :** {s['cdf_result']['texte']}")
        if s.get("do_compact"):
            comp=s["compacteurs"]
            st.markdown("**Compactage q4 :** "+" / ".join(f"{e}" for e,_ in comp["q4"]))
            st.markdown("**Compactage q3 :** "+" / ".join(f"{e}" for e,_ in comp["q3"]))

    st.markdown('<div class="note-box">⚠️ Rapport non officiel. GTR 2023 (IDRRIM/Cerema). À valider par un géotechnicien qualifié.</div>',unsafe_allow_html=True)

    st.divider()
    pdf_buf=build_pdf(pi,sondages)
    nom=f"Rapport_GTR2023_{pi['projet'][:20]}_{datetime.now().strftime('%Y%m%d')}.pdf"
    st.download_button("📥 Télécharger le rapport PDF complet",data=pdf_buf,file_name=nom,mime="application/pdf",use_container_width=True,type="primary")
    xl_buf = build_excel(pi, sondages)
    nom_xl = pi.get("projet","projet")[:20].replace(" ","_") or "projet"
    st.download_button("📊 Télécharger le rapport Excel",data=xl_buf,
        file_name=f"GTR2023_{nom_xl}_{datetime.now().strftime('%Y%m%d')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",use_container_width=True)
    st.divider()
    c1,c2,c3=st.columns(3)
    with c1:
        if st.button("➕ Ajouter un sondage",use_container_width=True): new_sondage(); st.rerun()
    with c2:
        proj_json = save_project()
        nom_projet = pi.get("projet","projet")[:30].replace(" ","_") or "projet"
        st.download_button("💾 Sauvegarder le projet", data=proj_json,
            file_name=f"GTR2023_{nom_projet}_{datetime.now().strftime('%Y%m%d')}.json",
            mime="application/json", use_container_width=True)
    with c3:
        if st.button("🔄 Nouveau projet",use_container_width=True): reset_all(); st.rerun()

# PIED DE PAGE
st.divider()
st.caption("Classification GTR2023 - S.M. SY / UIDT - Fascicules 1 et 2 (IDRRIM/Cerema, ed. 2024)")

# TODO: Synoptique visuel — tableau récapitulatif graphique avec couleurs par état/classe
# TODO: Profil en profondeur — représentation des couches par sondage avec épaisseurs
# TODO: Coupe transversale — coupe schématique entre sondages alignés
# TODO: Carte interactive — positionnement GPS des sondages sur fond de carte
# TODO: Graphiques comparatifs — diagrammes radar/bar des paramètres entre sondages
